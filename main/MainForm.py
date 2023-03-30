# import collections

from PyQt5.QtWidgets import QMainWindow
from PyQt5 import QtCore
import datetime

today = datetime.datetime.today().date()


class MainForm(QMainWindow):
    def __init__(self, parent):
        from PyQt5.QtCore import QDate
        from PyQt5.QtGui import QIcon
        from PyQt5.QtWidgets import QHeaderView, QFrame
        from PyQt5 import uic

        QMainWindow.__init__(self, parent)
        self.svr_rowspan: dict = {}
        self.svr_current_row: dict = {}

        self.ui = uic.loadUi("ui\MainForm.ui")
        self.ui.tableWgt1.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.ui.tableWgt1.horizontalHeader().setFixedHeight(40)
        self.ui.tableWgt1.horizontalHeader().setStyleSheet("QHeaderView::section {"
                                                           "background: rgb(240,255,255); "
                                                           "border: 1px solid rgb(203, 235, 255);"
                                                           "font-weight: 600;"
                                                           "color: rgb(95, 95, 95);}")

        self.ui.setWindowTitle(self.get_config("window_title"))
        self.ui.setWindowIcon(QIcon(self.get_config("icon")))
        self.ui.dateEdit1.setDate(QDate.currentDate())
        self.ui.show()

        self.ui.toolBtn1.clicked.connect(self.click_toolBtn1)
        self.ui.toolBtn2.clicked.connect(self.click_toolBtn2)
        self.ui.toolBtn3.clicked.connect(self.click_toolBtn3)
        self.ui.toolBtn4.clicked.connect(self.click_toolBtn4)
        self.ui.toolBtn5.clicked.connect(self.click_toolBtn5)
        self.ui.toolBtn6.clicked.connect(self.click_toolBtn6)
        self.ui.tableWgt1.clicked.connect(self.click_tableWgt1)
        self.ui.tableWgt1.installEventFilter(self)

        # resized = QtCore.pyqtSignal()
        # resized.connect(self.resize_check)

    @staticmethod
    def get_config(flag: str):
        from conf.config import Config
        return Config().conf_main_form(flag)

    def set_tableWgt1(self, date: datetime):
        """tableWgt1: Main table widget to present the result of the date"""
        from func.GetSvrInfo import GetSvrInfo
        from main.btn.SetMainForm import SetMainForm
        from PyQt5.QtWidgets import QTableWidgetItem

        self.ui.tableWgt1.clearContents()

        svr_list: dict = GetSvrInfo.get_svr_list()
        self.ui.tableWgt1.setRowCount(GetSvrInfo.get_row_count("TOTAL"))

        current_row = 0

        for i, id in enumerate(svr_list):
            s = SetMainForm(id, date)
            rowspan = GetSvrInfo.get_row_count(str(id))
            self.svr_rowspan[id] = rowspan
            self.svr_current_row[id] = current_row

            host_item = QTableWidgetItem(svr_list[id])
            self.ui.tableWgt1.setItem(current_row, 0, host_item)

            res_items = s.set_res()
            if res_items:
                self.ui.tableWgt1.setItem(current_row, 1, res_items[0])
                self.ui.tableWgt1.setItem(current_row, 2, res_items[1])

            res_rate_item = s.set_res_rate()
            if res_rate_item:
                self.ui.tableWgt1.setItem(current_row, 3, res_rate_item)

            letter_items, capacity_items = s.set_disk()
            disk_rate_items = s.set_disk_diff()
            for j in range(len(letter_items)):
                self.ui.tableWgt1.setItem(current_row + j, 4, letter_items[j])
                if len(capacity_items) > j:
                    self.ui.tableWgt1.setItem(current_row + j, 5, capacity_items[j])
                if len(disk_rate_items) > j:
                    self.ui.tableWgt1.setItem(current_row + j, 6, disk_rate_items[j])

            svc_item = s.set_svc()
            self.ui.tableWgt1.setItem(current_row, 7, svc_item)

            task_item = s.set_task()
            self.ui.tableWgt1.setItem(current_row, 8, task_item)

            wdef_item = s.set_wdef()
            self.ui.tableWgt1.setItem(current_row, 9, wdef_item)

            evt_item = s.set_evt()
            self.ui.tableWgt1.setItem(current_row, 10, evt_item)

            for col in range(self.ui.tableWgt1.columnCount()):
                if col not in self.get_config("use_separated_row"):
                    self.ui.tableWgt1.setSpan(current_row, col, rowspan, 1)
            current_row += rowspan

        self.ui.tableWgt1.resizeRowsToContents()
        # self.ui.tableWgt1.resizeColumnsToContents()

    def click_toolBtn1(self):
        """Refresh tableWgt1 (date: today)"""
        self.set_tableWgt1(today)

    def click_toolBtn2(self):
        """Insert a new server"""
        from main.btn.InsertSvr import InsertSvr
        InsertSvr(self)

    def click_tableWgt1(self) -> int:
        return self.ui.tableWgt1.currentRow()

    def click_toolBtn3(self):
        """Delete the server"""
        from main.btn.DeleteSvr import DeleteSvr_New
        # DeleteSvr().delete(self.click_tableWgt1())
        DeleteSvr_New(self)

    def click_toolBtn4(self):
        """Refresh tableWgt1 (date: selected date)"""
        date = datetime.datetime.strptime(self.ui.dateEdit1.date().toString('yyyy-MM-dd'), '%Y-%m-%d').date()
        self.set_tableWgt1(date)

    def click_toolBtn5(self):
        """To manage syscode, gbcode"""
        from main.btn.ManageCode import ManageCode
        ManageCode(self)

    def click_toolBtn6(self):
        """Download Excel file"""
        self.qtablewidget_to_excel()

    def get_svrid(self, current_row: int) -> int:
        try:
            return next(key for key, value in self.svr_current_row.items() if value == current_row)
        except StopIteration:
            val = 0
            for value in self.svr_current_row.values():
                if value < current_row:
                    val = max(val, value)
            return next(key for key, value in self.svr_current_row.items() if value == val)

    '''def resizeEvent(self, event):
        self.resized.emit()
        return super(MainForm, self).resizeEvent(event)

    def resize_check(self):
        self.ui.width = self.ui.centralWidget.frameSize().width() - self.ui.verticalLayout.maximumSize().width()'''

    def eventFilter(self, source, event):
        """To copy and paste selected items in tableWgt1"""
        from PyQt5.QtCore import QEvent
        from PyQt5.QtGui import QKeySequence

        if source == self.ui.tableWgt1 and event.type() == QEvent.KeyPress and event == QKeySequence.Copy and not event.isAutoRepeat():
            # copy_string = ""
            # for i in self.ui.tableWgt1.selectedIndexes():
            #     copy_string += "\n" if i.column() == 0 and copy_string else ""
            #     copy_string += i.data() if i.data() else ""
            #     copy_string += "\t"

            datadict = {}
            for i, id in enumerate(self.svr_rowspan):
                datadict[id] = {}
                for j in range(self.svr_rowspan[id]):
                    datadict[id][j] = []

            for i in self.ui.tableWgt1.selectedIndexes():
                id = self.get_svrid(i.row())
                datadict[id][i.row() - self.svr_current_row[id]].append(i.data() if i.data() else "")

            copy_string = self.datadict_to_html(datadict)

            from PyQt5.QtWidgets import QApplication
            QApplication.clipboard().setText(copy_string)
            return True
        return super(QMainWindow, self).eventFilter(source, event)

    def datadict_to_html(self, datadict: dict) -> str:
        copy_string: str = self.get_config("copy_string")
        copy_html: list = ["<html>\n", "\t<table>\n", copy_string]
        for key, value in datadict.items():
            rowspan = self.svr_rowspan[key]
            for k, v in value.items():
                copy_html.append("\t\t<tr>\n")
                for i, val in enumerate(v):
                    copy_html.append(f"\t\t\t<td rowspan={rowspan}>" if i not in self.get_config(
                        "use_separated_row") and k == 0 else "\t\t\t<td>")
                    copy_html.append(val)
                    copy_html.append("</td>\n")
                copy_html.append("\t\t</tr>\n")
        copy_html.append("\t</table>\n")
        copy_html.append("</html>\n")
        return "".join(copy_html)

    def qtablewidget_to_excel(self):
        import openpyxl
        from openpyxl.styles import PatternFill, Alignment
        from datetime import datetime

        if self.get_config("license") == "JAEMU":
            today = datetime.today().strftime('%Y%m%d')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = today

            alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
            names = ['호스트명', 'CPU', '메모리', '메모리 증감율', '디스크', '디스크 용량', '디스크 증감율', '서비스', '작업스케줄러', '윈도우디펜더', '이벤트']

            # 맨 윗 줄 세팅
            for i in range(0, 11):
                ws[f'{alphabet[i]}1'] = names[i]
                ws.column_dimensions[alphabet[i]].width = 16

            # 데이터 세팅
            tableWgtRowCnt = self.ui.tableWgt1.rowCount()
            tableWgtColCnt = self.ui.tableWgt1.columnCount()
            nowServer = ""

            for row in range(0, tableWgtRowCnt):
                excelRow = row + 2
                # 호스트명 세팅
                if self.ui.tableWgt1.item(row, 0) is not None:
                    nowServer = self.ui.tableWgt1.item(row, 0).text()
                ws[f'A{excelRow}'] = nowServer

                for col in range(1, tableWgtColCnt):
                    item = self.ui.tableWgt1.item(row, col)
                    txt = '' if item is None else item.text()
                    ws[f'{alphabet[col]}{excelRow}'] = txt

            wb.save(r'C:\DailyCheck\file\재무파트 서버 데일리 체크_' + today + '.xlsx')

            print("DONE")
